"""
Report Generation Service for ASECNA Stock Management

Generates weekly/monthly reports with stock movements, withdrawals, and AI analytics.
Supports Excel and CSV export formats.
"""

import pandas as pd
from datetime import datetime, timedelta
from typing import Optional, Dict, List
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ReportService:
    """
    Service for generating comprehensive stock reports.
    
    Features:
    - Weekly/Monthly reports
    - Multiple sheets (Summary, By Country, By Category, etc.)
    - Excel (.xlsx) and CSV export
    - Scheduled automatic generation
    """
    
    def __init__(self, db_session=None):
        self.db = db_session
    
    def generate_weekly_report(
        self,
        start_date: Optional[datetime] = None,
        end_date: Optional[datetime] = None
    ) -> Dict:
        """
        Generate weekly report with all analytics.
        
        Args:
            start_date: Report start date (default: 7 days ago)
            end_date: Report end date (default: today)
            
        Returns:
            Dictionary with report data and metadata
        """
        if not start_date:
            start_date = datetime.now() - timedelta(days=7)
        if not end_date:
            end_date = datetime.now()
        
        logger.info(f"Generating weekly report from {start_date} to {end_date}")
        
        # Collect all report data
        report_data = {
            'metadata': {
                'report_type': 'weekly',
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'generated_at': datetime.now().isoformat()
            },
            'summary': self._generate_summary(start_date, end_date),
            'by_country': self._generate_country_breakdown(start_date, end_date),
            'by_category': self._generate_category_breakdown(start_date, end_date),
            'withdrawals': self._generate_withdrawals_data(start_date, end_date),
            'ai_accuracy': self._generate_ai_metrics(start_date, end_date),
            'alerts': self._generate_alerts_data(start_date, end_date)
        }
        
        return report_data
    
    def _generate_summary(self, start_date: datetime, end_date: datetime) -> Dict:
        """Generate summary statistics"""
        # TODO: Replace with actual database queries
        return {
            'total_items_start': 1420,
            'total_items_end': 1385,
            'total_movements': 156,
            'entries': 45,
            'exits': 111,
            'withdrawals_processed': 23,
            'critical_alerts': 3,
            'ai_detections': 892
        }
    
    def _generate_country_breakdown(self, start_date: datetime, end_date: datetime) -> List[Dict]:
        """Generate withdrawals breakdown by country"""
        # Mock data - replace with actual queries
        countries = [
            {'country': 'Côte d\'Ivoire', 'withdrawals': 8, 'quantity': 125},
            {'country': 'Mali', 'withdrawals': 5, 'quantity': 78},
            {'country': 'Burkina Faso', 'withdrawals': 4, 'quantity': 62},
            {'country': 'Sénégal', 'withdrawals': 3, 'quantity': 45},
            {'country': 'Gabon', 'withdrawals': 2, 'quantity': 28},
            {'country': 'Cameroun', 'withdrawals': 1, 'quantity': 15}
        ]
        return countries
    
    def _generate_category_breakdown(self, start_date: datetime, end_date: datetime) -> List[Dict]:
        """Generate stock distribution by category"""
        categories = [
            {'category': 'RAM', 'quantity': 450, 'withdrawals': 85},
            {'category': 'Desktop Computers', 'quantity': 320, 'withdrawals': 42},
            {'category': 'Laptops', 'quantity': 215, 'withdrawals': 28},
            {'category': 'Air Conditioners', 'quantity': 180, 'withdrawals': 15},
            {'category': 'Printers', 'quantity': 120, 'withdrawals': 12},
            {'category': 'Network Equipment', 'quantity': 100, 'withdrawals': 8}
        ]
        return categories
    
    def _generate_withdrawals_data(self, start_date: datetime, end_date: datetime) -> List[Dict]:
        """Generate detailed withdrawals data"""
        withdrawals = [
            {
                'id': 'WD-0012',
                'date': '2026-02-13',
                'country': 'Côte d\'Ivoire',
                'equipment': 'RAM DDR4 8GB',
                'quantity': 50,
                'status': 'Completed'
            },
            {
                'id': 'WD-0013',
                'date': '2026-02-12',
                'country': 'Mali',
                'equipment': 'Desktop Computer',
                'quantity': 10,
                'status': 'In Transit'
            }
        ]
        return withdrawals
    
    def _generate_ai_metrics(self, start_date: datetime, end_date: datetime) -> Dict:
        """Generate AI detection metrics"""
        return {
            'total_detections': 892,
            'average_confidence': 0.91,
            'counting_accuracy': 0.89,
            'false_positives': 12,
            'missed_detections': 8
        }
    
    def _generate_alerts_data(self, start_date: datetime, end_date: datetime) -> List[Dict]:
        """Generate alerts/anomalies data"""
        alerts = [
            {
                'date': '2026-02-11',
                'type': 'Stock Discrepancy',
                'severity': 'Medium',
                'description': 'RAM count mismatch: Declared 100, Detected 98',
                'resolved': True
            },
            {
                'date': '2026-02-10',
                'type': 'Unauthorized Movement',
                'severity': 'High',
                'description': 'Equipment detected in restricted zone',
                'resolved': False
            }
        ]
        return alerts
    
    def export_to_excel(
        self,
        report_data: Dict,
        filename: Optional[str] = None
    ) -> bytes:
        """
        Export report to Excel format with multiple sheets.
        
        Args:
            report_data: Report data dictionary
            filename: Optional filename (auto-generated if not provided)
            
        Returns:
            Excel file as bytes
        """
        if not filename:
            date_str = datetime.now().strftime('%Y-%m-%d')
            filename = f'ASECNA_Weekly_Report_{date_str}.xlsx'
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet('Summary')
        self._create_summary_sheet(ws_summary, report_data)
        
        # 2. By Country Sheet
        ws_country = wb.create_sheet('By Country')
        self._create_country_sheet(ws_country, report_data['by_country'])
        
        # 3. By Category Sheet
        ws_category = wb.create_sheet('By Category')
        self._create_category_sheet(ws_category, report_data['by_category'])
        
        # 4. Withdrawals Sheet
        ws_withdrawals = wb.create_sheet('Withdrawals')
        self._create_withdrawals_sheet(ws_withdrawals, report_data['withdrawals'])
        
        # 5. Alerts Sheet
        ws_alerts = wb.create_sheet('Alerts')
        self._create_alerts_sheet(ws_alerts, report_data['alerts'])
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Excel report generated: {filename}")
        return excel_buffer.getvalue()
    
    def _create_summary_sheet(self, ws, report_data):
        """Create summary sheet with key metrics"""
        # Header
        ws['A1'] = 'ASECNA Stock Management - Weekly Report'
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A2'] = f"Period: {report_data['metadata']['start_date']} to {report_data['metadata']['end_date']}"
        ws['A3'] = f"Generated: {report_data['metadata']['generated_at']}"
        
        # Metrics
        summary = report_data['summary']
        row = 5
        metrics = [
            ('Total Items (Start)', summary['total_items_start']),
            ('Total Items (End)', summary['total_items_end']),
            ('Total Movements', summary['total_movements']),
            ('Entries', summary['entries']),
            ('Exits', summary['exits']),
            ('Withdrawals Processed', summary['withdrawals_processed']),
            ('Critical Alerts', summary['critical_alerts']),
            ('AI Detections', summary['ai_detections'])
        ]
        
        for label, value in metrics:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Auto-adjust column width
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _create_country_sheet(self, ws, data):
        """Create country breakdown sheet"""
        headers = ['Country', 'Withdrawals', 'Total Quantity']
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='137fec', end_color='137fec', fill_type='solid')
        
        # Data rows
        for item in data:
            ws.append([item['country'], item['withdrawals'], item['quantity']])
    
    def _create_category_sheet(self, ws, data):
        """Create category breakdown sheet"""
        headers = ['Category', 'Current Stock', 'Withdrawals']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        
        for item in data:
            ws.append([item['category'], item['quantity'], item['withdrawals']])
    
    def _create_withdrawals_sheet(self, ws, data):
        """Create withdrawals detail sheet"""
        headers = ['Request ID', 'Date', 'Country', 'Equipment', 'Quantity', 'Status']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
        
        for item in data:
            ws.append([
                item['id'],
                item['date'],
                item['country'],
                item['equipment'],
                item['quantity'],
                item['status']
            ])
    
    def _create_alerts_sheet(self, ws, data):
        """Create alerts/anomalies sheet"""
        headers = ['Date', 'Type', 'Severity', 'Description', 'Resolved']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
        
        for item in data:
            ws.append([
                item['date'],
                item['type'],
                item['severity'],
                item['description'],
                'Yes' if item['resolved'] else 'No'
            ])
    
    def export_to_csv(
        self,
        report_data: Dict,
        sheet_name: str = 'summary'
    ) -> bytes:
        """
        Export specific sheet to CSV format.
        
        Args:
            report_data: Report data dictionary
            sheet_name: Which sheet to export
            
        Returns:
            CSV file as bytes
        """
        if sheet_name == 'summary':
            df = pd.DataFrame([report_data['summary']])
        elif sheet_name == 'by_country':
            df = pd.DataFrame(report_data['by_country'])
        elif sheet_name == 'by_category':
            df = pd.DataFrame(report_data['by_category'])
        elif sheet_name == 'withdrawals':
            df = pd.DataFrame(report_data['withdrawals'])
        elif sheet_name == 'alerts':
            df = pd.DataFrame(report_data['alerts'])
        else:
            raise ValueError(f"Unknown sheet name: {sheet_name}")
        
        csv_buffer = io.StringIO()
        df.to_csv(csv_buffer, index=False)
        csv_buffer.seek(0)
        
        return csv_buffer.getvalue().encode('utf-8')


# Singleton instance
_report_service = None


def get_report_service(db_session=None):
    """Get or create the global report service instance"""
    global _report_service
    if _report_service is None:
        _report_service = ReportService(db_session)
    return _report_service
